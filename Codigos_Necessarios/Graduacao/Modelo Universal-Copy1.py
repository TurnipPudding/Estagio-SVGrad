# -*- coding: utf-8 -*-
"""
## Import das bibliotecas
"""

# Imports das bibliotecas e funções utilizadas
import pandas as pd # Leitura de dados
import time # Cálculo de tempo
from mip import Model, xsum, minimize, BINARY, INTEGER, CONTINUOUS, OptimizationStatus # Biblioteca com linguagem de modelagem
import os
import sys # Implementação de saídas de erro
import traceback # Implementação de inspeção para auxiliar nas saídas de erro.
# Métodos e funções para auxiliar na criação de planilhas mais elaboradas
from openpyxl import Workbook
from openpyxl.formatting.rule import DataBar, FormatObject, Rule

"""
## Leitura dos dados
Etapa responsável por carregar os dados das planilhas de cada departamento e das salas, além de extrair listas auxiliares para análise.
"""

# Caminho do arquivo Excel passado como argumento
file_path = sys.argv[1]

# Lê o arquivo principal com todas as planilhas (salas e disciplinas)
df_completo = pd.ExcelFile(sys.argv[1])
# Extrai os nomes das planilhas presentes no arquivo
sheet_names = df_completo.sheet_names

# Lê e mescla os dados das aulas de todos os departamentos em um único DataFrame
df = pd.read_excel(file_path, sheet_name=sheet_names[1:])
df = pd.concat(df.values(), ignore_index=True)
# print(df)  # Para depuração

# Lê os dados das salas
salas = pd.read_excel(file_path, sheet_name=sheet_names[0])
# print(salas)  # Para depuração

# Lista da capacidade de cada sala
cap_s = salas['Lugares'].tolist()
# print(cap_s)  # Para depuração

# Lista do tamanho de cada disciplina (número de inscritos)
tam_t = df['Vagas por disciplina'].tolist()
# print(tam_t)  # Para depuração

print('\nBase de Dados lida.')


"""
## Tratamento dos Dados
Etapa responsável por padronizar horários, tratar células irregulares e garantir consistência dos dados para análise posterior.
"""

# Função para converter horário no formato 'HH:MM' para valor decimal em horas
def horario_para_decimal(horario):
    # Se houver horário no formato '20h40', converte para '20:40'
    if 'h' in horario:
        horario = horario.replace('h',':')
    # Separa horas e minutos
    horas, minutos = map(int, horario.split(':'))
    # Retorna valor decimal (ex: 20:40 -> 20.67)
    return horas + minutos / 60

# Função para processar célula no formato 'Dia - HH:MM/HH:MM'
def processar_horario(celula):
    # Verifica se a célula possui horário definido
    if isinstance(celula, str) and "-" in celula:
        # Remove espaços extras
        celula = str(celula).replace(' ', '')
        # Separa dia e horários
        dia, horarios = celula.split('-')
        inicio, fim = horarios.split('/')
        # Converte horários para decimal
        start_a = horario_para_decimal(inicio)
        end_a = horario_para_decimal(fim)
        return dia, start_a, end_a
    else:
        # Célula irregular ou vazia
        return 0, 0, 0

# Lista das colunas de horários a serem processadas
colunas_horarios = ['Horário 1', 'Horário 2', 'Horário 3', 'Horário 4']
result = []
# Processa cada coluna de horários e traduz para formato padronizado
for coluna in colunas_horarios:
    resultados = df[coluna].apply(processar_horario).to_list()
    result.extend(resultados)

# Cria DataFrame com dados padronizados de todas as aulas
A = pd.DataFrame(result, columns=['Dia', 'start_a', 'end_a'])

# Salva colunas do DataFrame em listas separadas para facilitar análise
dia_a = A['Dia'].to_list()
start_a = A['start_a'].to_list()
end_a = A['end_a'].to_list()
# print(A)
# print(len(A))
# print(start_a)

# Preenche células vazias na coluna 'Sala' com valor 0
for s in range(len(df['Sala'])):
    if pd.isna(df.loc[s, 'Sala']):
        df.loc[s, 'Sala'] = '0'

# Preenche células vazias na coluna 'Turma' com valor 1
for d in range(len(tam_t)):
    if pd.isna(df.loc[d, 'Turma']):
        df.loc[d, 'Turma'] = 1

"""## Dados de Entrada

Bloco responsável por criar listas e dicionários que representam os índices das turmas/disciplina, salas, laboratórios, cursos e suas relações. Essas estruturas são fundamentais para a modelagem e análise dos dados de alocação de aulas e salas."""

# Lista de índices de cada turma/disciplina (0, 1, ..., n-1)
T = range(len(df['Disciplina (código)']))
# Exemplo de uso: T[38] retorna o índice da disciplina na posição 38
# Exemplo: df.loc[T[38], 'Disciplina (código)'] retorna o código da disciplina

# Lista de índices de cada sala (0, 1, ..., m-1)
S = range(len(salas['Sala']))
# Exemplo de uso: salas.loc[S[2], 'Sala'] retorna o nome da sala na posição 2

# Lista binária indicando se cada sala é laboratório (1 = sim, 0 = não)
sigma_s = [1 if salas.loc[s, 'Laboratório'] == 'Sim' else 0 for s in S]
# sigma_s[s] = 1 se sala s é laboratório

# Lista binária indicando se cada turma/disciplina precisa de laboratório (1 = sim, 0 = não)
tal_t = [0 if df.loc[t, 'Utilizará laboratório? (sim ou não)'] == 'Não' else 1 for t in T]
# tal_t[t] = 1 se turma t precisa de laboratório

# Salva os comprimentos das listas principais para evitar recomputação
lenT = len(T)   # Número de turmas/disciplina
lenA = len(A)   # Número total de aulas
lenS = len(S)   # Número de salas

# Lista dos cursos/currículos do ICMC
curriculos = ['BMACC', 'BMA', 'LMA', 'MAT-NG', 'BECD', 'BCC', 'BSI', 'BCDados']
lenC = len(curriculos)  # Número de cursos

# Dicionário Y_tc: (turma, curso) -> 1 se a turma é ministrada para o curso, 0 caso contrário
Y_tc = {(t, c): 0 for t in range(lenT) for c in range(lenC)}
for t in range(lenT):
    celula = df.loc[t, 'Curso(s)']
    # Se há mais de um curso, separa por vírgula e marca todos os cursos presentes
    if ',' in celula:
        for c in celula.split(', '):
            if c in curriculos:
                Y_tc[(t, curriculos.index(c))] = 1
    else:
        # Se há apenas um curso, marca se estiver na lista
        if celula in curriculos:
            Y_tc[(t, curriculos.index(celula))] = 1
# Y_tc[(t, c)] = 1 se turma t é ministrada para curso c


"""## Dados de Preprocessamento

Bloco responsável por criar listas e dicionários auxiliares para análise e restrições do problema de alocação. Aqui são definidos agrupamentos de aulas, restrições de capacidade, conflitos de horário, uso de espaço e distâncias entre salas."""

# A_t: lista de listas, cada sublista contém os índices das aulas de uma turma/disciplina
# Exemplo: A_t[0] = [0, 1, 2] significa que as aulas 0, 1 e 2 pertencem à turma 0
A_t = []
for t in range(lenT):
    # Para cada turma, calcula os índices das aulas associadas
    A_t.append([t + (i * lenT) for i in range(int(lenA/lenT))])
# A_t é útil para mapear aulas por turma

# A_tt: igual ao A_t, mas exclui aulas sem horário definido (start_a[a] == 0)
A_tt = [[a for a in A_t[t] if start_a[a] != 0] for t in range(lenT)]
# A_tt permite filtrar apenas aulas válidas para análise

# A_s: lista de listas, cada sublista contém índices de aulas seguidas de uma mesma turma no mesmo dia
# Usada para garantir que aulas consecutivas fiquem na mesma sala
A_s = []
for t in A_t:
    # Verifica se as duas primeiras aulas da turma são no mesmo dia e estão definidas
    if dia_a[t[0]] == dia_a[t[1]] and dia_a[t[0]] != 0:
        # Verifica se o intervalo entre as aulas é compatível (não há sobreposição indevida)
        if end_a[t[0]] + 2 + 10/60 + 20/60 >= start_a[t[1]]:
            seguido = [x for x in t if start_a[x] != 0]
            if len(seguido) > 1:
                A_s.append(seguido)
# A_s é usada para restrições de alocação de aulas consecutivas

# A_c: lista de listas, cada sublista contém índices das aulas ministradas para cada curso
A_c = []
for c in range(lenC):
    # Para cada curso, inclui todas as aulas das turmas que o ministram
    A_c.append([t + (i * lenT) for i in range(int(lenA/lenT)) for t in range(lenT) if Y_tc[t, c] == 1])
# A_c[c] contém todas as aulas do curso c

# eta_as: dicionário (aula, sala) -> 1 se sala comporta a aula, 0 caso contrário
# Usa a capacidade da sala e o tamanho da turma
eta_as = {(a, s): 1 if tam_t[int((a % lenT))] <= cap_s[s] else 0 for a in range(lenA) for s in range(lenS)}
# eta_as[(a, s)] = 1 se sala s comporta aula a

# theta_aal: dicionário (aula, aula) -> 1 se há conflito de horário entre as aulas, 0 caso contrário
# Conflito ocorre se as aulas são no mesmo dia e os horários se sobrepõem
theta_aal = {(a, al): 1 if (dia_a[a] == dia_a[al] and (start_a[a] < end_a[al] and start_a[al] < end_a[a])) else 0 for a in range(lenA) for al in range(lenA)}
# theta_aal[(a, al)] = 1 se há conflito de horário entre aula a e aula al

# uso_as: dicionário (aula, sala) -> percentual de espaço vazio na sala ao alocar a aula
# Quanto menor o valor, melhor o aproveitamento da sala
uso_as = {(a, s): 100 * (1 - (tam_t[int((a % lenT))]/cap_s[s])) for a in range(lenA) for s in range(lenS)}
# uso_as[(a, s)] = percentual de espaço vazio

# dis: dicionário (sala, sala) -> distância arbitrária entre duas salas
# Usado para restrições de deslocamento entre salas
dis = {(s, sl): salas.loc[s, sl] for s in range(len(salas)) for sl in salas.columns[3:-1]}
# dis[(s, sl)] = distância entre sala s e sala sl

# Lista de salas preferencialmente vazias
pref = salas['Preferencialmente Vazia'].tolist()


"""## Dados para fixar os laboratórios e aulas com sala definida, como LEM

Bloco responsável por definir restrições específicas de alocação: aulas que exigem laboratório, aulas com sala fixa, e proibições de horários/salas. Essas estruturas garantem que certas aulas sejam obrigatoriamente alocadas em laboratórios ou salas específicas, e que restrições de uso sejam respeitadas pelo modelo.
"""

# labs: lista de índices das turmas/disciplina que precisam de laboratório (tal_t = 1)
labs = [t for t in T if tal_t[t] == 1]
# salas_labs: lista de índices das salas que são laboratórios
salas_labs = [i for i in range(len(salas['Laboratório'])) if salas.loc[i, 'Laboratório'] == "Sim"]

# ind_labs: lista de listas, cada sublista indica quais aulas de cada turma/disciplina devem ser em laboratório
# Exemplo: ind_labs[0] = [1,2] significa que as aulas 1 e 2 da turma labs[0] devem ser em laboratório
ind_labs = []
for l in labs:
    # Extrai os índices das aulas que devem ser em laboratório, removendo o texto 'Sim' e convertendo para inteiro
    valores = (df.loc[l, 'Utilizará laboratório? (sim ou não)'].replace(' ', '')).split(',')
    if "Sim" in valores:
        valores.remove("Sim")
    elif "sim" in valores:
        valores.remove("sim")
    ind_labs.append([int(item) for item in valores])

# aula_labs: lista de índices de todas as aulas que devem ser ministradas em laboratório
# Calcula o índice global da aula usando a relação entre turma e posição da aula
aula_labs = [(labs[t] + (lenT * (i-1))) for t in range(len(labs)) for i in ind_labs[t]]

# lab_tal: lista binária, cada posição indica se a aula é de laboratório (1) ou não (0)
lab_tal = [0 for _ in range(lenA)]
for a in aula_labs:
    lab_tal[a] = 1

# sala_fixa: lista de nomes de salas fixadas para cada aula; '0' indica sem sala fixa
# Se a célula de sala tem mais de um valor, seleciona o correto conforme o horário
sala_fixa = []
for a in range(lenA):
    sala_valor = str((df.loc[a % lenT, 'Sala']))
    if ', ' in sala_valor:
        if not pd.isna(df.loc[int(a / lenT), 'Horário ' + str(int(a / lenT) + 1)]):
            if len(sala_valor.split(', ')) >= (int(a / lenT) + 1):
                sala_fixa.append(sala_valor.split(', ')[int(a / lenT)])
            else:
                sala_fixa.append('0')
    elif not pd.isna(df.loc[int(a / lenT), 'Horário ' + str(int(a / lenT) + 1)]):
        sala_fixa.append(sala_valor)
    else:
        sala_fixa.append('0')

# Aplica restrição de sala fixa: zera todas as possibilidades de sala para a aula, exceto a sala fixada
for aula in range(lenA):
    if sala_fixa[aula] != '0':
        fixada = salas['Sala'].tolist().index(sala_fixa[aula])
        for sala in range(lenS):
            eta_as[aula, sala] = 0
        eta_as[aula, fixada] = 1

# Aplica restrição de laboratório: aulas de laboratório só podem ser alocadas em salas de laboratório, e vice-versa
for aula in range(lenA):
    if lab_tal[aula] == 1:
        for sala in range(lenS):
            if sala not in salas_labs:
                eta_as[aula, sala] = 0
    else:
        for sala in range(lenS):
            if sala in salas_labs:
                eta_as[aula, sala] = 0

# sala_proibida: dicionário que armazena restrições de salas proibidas para cada aula
# Se a célula de proibição tem mais de um valor, aplica restrição para todas as salas listadas
sala_proibida = {}
for a in range(lenA):
    cell = df.loc[int(a % lenT), 'Proibir Horário ' + str(int(a / lenT) + 1)]
    if not pd.isna(cell):
        cell = str(cell)
        if ',' in cell:
            salas_proibidas = cell.split(', ')
            sala_proibida[a] = salas_proibidas
            for sala in salas_proibidas:
                s = salas[salas['Sala'] == sala].index[0]
                eta_as[a, s] = 0
        else:
            sala_proibida[a] = cell
            s = salas[salas['Sala'] == cell].index[0]
            eta_as[a, s] = 0

# seguidas: lista que indica se a turma tem aulas seguidas (2) ou não (1); usada para penalizar trocas de sala
seguidas = [1 for _ in range(lenT)]
for t in A_s:
    seguidas[int(t[0] % lenT)] = 2


"""## Modelo Principal"""

# Criação do Modelo.
model = Model("Alocação de aulas", solver_name="CBC")

# Variáveis de decisão.
# x_as é uma variável binário que ganha o valor 1 se a aula 'a' é alocada à sala 's', e ganha o valor 0 no caso contrário.
# Ela está na forma de um dicionário na forma de matriz para facilitar o código, simulando uma matriz de variáveis x00, x01, x02,..., xnm.
x_as = {(a, s): model.add_var(var_type=BINARY) for a in range(lenA) for s in range(lenS)}
# A variável peso_x é o quanto a variável x_as afeta o modelo, ou seja, o valor de peso_x descreve a importância de x_as no modelo.
peso_x = int(sys.argv[2])


# y_t é uma variável inteira que contabiliza o número de trocas de sala de uma turma/disciplina.
# Ex: Se há uma disciplina com três aulas na semana, todas em salas diferentes, então y_t teria o valor 2 (já que houveram duas trocas).
# Ela está na forma de uma lista para facilitar o código, simulando um vetor de variáveis y0, y1, y2,..., yt.
y_t = [model.add_var(var_type=INTEGER, lb=0) for t in range(len(T))]
# A variável peso_y é o quanto a variável y_t afeta o modelo, ou seja, o valor de peso_y descreve a importância de y_t no modelo.
peso_y = int(sys.argv[3])

# c_st é uma variável binária que ganha o valor 1 se a sala 's' é utilizada pela turma/disciplina 't', e ganha o valor 0 no caso contrário.
# Como a variável x_as, o uso do dicionário na forma de matriz é para facilitar o código, simulando uma matriz de variáveis c00, c01,..., cmt.
# Caso esta variável esteja sendo usada, a interpretação de y_t é alterada para ser o número de salas usadas pela turma/disciplina 't'.
c_st = {(s, t): model.add_var(var_type=BINARY) for s in range(lenS) for t in range(lenT)}

obj = peso_x * xsum(uso_as[a,s] * x_as[a,s] for a in range(lenA) for s in range(lenS)) + peso_y*xsum(seguidas[t] * y_t[t] for t in range(lenT))

if sys.argv[4]:
    # w_cs é uma variável binária que ganha o valor 1 se o curso 'c' tem ao menos uma aula na sala 's', e ganha 0 caso contrário.
    # Ela está na forma de um dicionário na forma de matriz para facilitar o código, simulando uma matriz de variáveis w00, w01,...,wkm
    w_cs = {(c, s): model.add_var(var_type=BINARY) for c in range(lenC) for s in range(lenS)}
    
    
    # v_cssl é uma variável binária que ganha o valor 1 se o curso 'c' tem ao menos uma aula na sala 's' e na sala 'sl', e ganha 0 caso contrário.
    # Ela está na forma de um dicionário na forma de uma matriz tridimensional para facilitar o código, simulando uma matriz de variáveis
    # v000, v001,..., v00m, v010, v011,..., vkmm
    v_cssl = {(c, s, sl): model.add_var(var_type=BINARY) for c in range(lenC) for s in range(lenS) for sl in range(lenS) if s != sl}
    # A variável peso_v é o quanto a variável v_cssl afeta o modelo, ou seja, o valor de peso_v descreve a importância de v_cssl no modelo.
    peso_v = int(sys.argv[4])

    obj = obj + peso_v * xsum(dis[s,sl] * v_cssl[c,s,sl] for c in range(lenC) for s in range(lenS) for sl in range(lenS) if s != sl)
if sys.argv[5] and sys.argv[6]:
    # Variável de superlotação.
    # z_as é uma variável binária que ganha o valor 1 se a aula 'a', ao ser alocada à sala 's', ultrapassa uma certa quantia do espaço.
    # Ex: Se o fator de superlotação for 0.85, então a variável ganhará o valor 1 se a aula 'a' ocupar mais de 85% do espaço disponível da sala 's'.
    # Essa variável possui dois propósitos: acomodar melhor uma aula para melhorar o conforto dos alunos, e considerar o possível
    # aumento de inscritos na segunda interação de matrícula e no período de requerimento.
    z_as = {(a, s): model.add_var(var_type=BINARY) for a in range(lenA) for s in range(lenS)}
    # A variável peso_z é o quanto a variável z_as afeta o modelo, ou seja, o valor de peso_z descreve a importância de z_as no modelo.
    peso_z = int(sys.argv[5])
    alpha = float(sys.argv[6])

    obj = obj + peso_z * xsum(z_as[a, s] for a in range(lenA) for s in range(lenS))
if sys.argv[7]:
    peso_pref = int(sys.argv[7])

    obj = obj + peso_pref * xsum(pref[s] * x_as[a,s] for a in range(lenA) for s in range(lenS))


# Função Objetivo.
# Queremos minimizar o espaço vazio das salas, descrito pela somatória de (uso_as * x_as), que contabiliza o uso da sala 's' pela aula 'a'.
# Além disso, queremos também minimizar o número de trocas de sala, descrito pela somatória de y_t, que contabiliza
# quantas trocas de sala foram feitas pela turma/disciplina 't'.
# Caso a variável c_st esteja sendo usada, por ela ser uma variável de "verificação", ela não precisa ser contabilizada na função objetivo.
# Caso a variável c_st esteja sendo usada, a interpretação de y_t muda, mas sua influência na função objetivo não é alterada.
# Pelo mesmo motivo de c_st não estar contabilizada na função objetivo, a variável w_cs também não está.

model.objective = minimize(obj)

# Restrições

# (3.3) - A sala precisa ser capaz de receber todos os inscritos da aula.
# Essa restrição garante que uma aula 'a' só pode ser alocada em uma sala 's' que consegue suportá-la, e isso é verificado
# utilizando a variável eta_as, que possui o valor 1 se 's' suporta 'a', e 0 caso contrário.
# Para toda aula 'a'.
for a in range(lenA):
    if start_a[a] == 0 or tam_t[a % lenT] == 0:
        model += xsum(x_as[a,s] for s in range(lenS)) == 0
    if start_a[a] != 0 and tam_t[a % lenT] != 0:
        model += xsum(x_as[a,s] for s in range(lenS)) == 1
    # Para toda aula 's'.
    for s in range(lenS):
        # Garanto que a aula 'a' deva ser alocada em uma sala 's' que a suporta.
        model += x_as[a,s] <= eta_as[a,s]

# (3.4) - Aulas de mesmo horário não podem estar alocadas na mesma sala
# Essa restrição garante que as aulas com conflito de horário sejam alocadas em salas diferentes.
# Para cada sala 'a'.
for a in range(lenA):
    # Para cada sala 'al'.
    for al in range(lenA):
        # Verifico se as aulas 'a' e 'al' possuem conflito de horário.
        if theta_aal[a, al] == 1 and a != al:
            # Se elas são aulas diferentes e possuem conflito de horário, 
            # então garanto que as duas não possam ser alocadas para a mesma sala.
            for s in range(lenS):
                model += x_as[a,s] + x_as[al,s] <= 1
            # Caso ambas as disciplinas sejam de laboratório, também é necessário considerar o conflito de horários entre as
            # salas de laboratório que podem ser conjuntas, isto é, se uma aula for alocada na sala 6-303, ela terá que exibir
            # conflito com uma aula que acontece no mesmo horário nas salas 6-303/6-304
            if lab_tal[a] == 1 and lab_tal[al] == 1:
                # Restringe alocação envolvendo salas conjuntas.
                # Se uma aula está em 6-303 ou 6-304, a outra não pode estar em 6-303/6-304.
                model += x_as[a, salas['Sala'].to_list().index('6-303/6-304')] + \
                x_as[al, salas['Sala'].to_list().index('6-303')] <= 1
                model += x_as[a, salas['Sala'].to_list().index('6-303/6-304')] + \
                x_as[al, salas['Sala'].to_list().index('6-304')] <= 1
                model += x_as[al, salas['Sala'].to_list().index('6-303/6-304')] + \
                x_as[a, salas['Sala'].to_list().index('6-303')] <= 1
                model += x_as[al, salas['Sala'].to_list().index('6-303/6-304')] + \
                x_as[a, salas['Sala'].to_list().index('6-304')] <= 1

                # Restringe alocação envolvendo salas conjuntas.
                # Se uma aula está em 6-305 ou 6-306, a outra não pode estar em 6-305/6-306.
                model += x_as[a, salas['Sala'].to_list().index('6-305/6-306')] + \
                x_as[al, salas['Sala'].to_list().index('6-305')] <= 1
                model += x_as[a, salas['Sala'].to_list().index('6-305/6-306')] + \
                x_as[al, salas['Sala'].to_list().index('6-306')] <= 1
                model += x_as[al, salas['Sala'].to_list().index('6-305/6-306')] + \
                x_as[a, salas['Sala'].to_list().index('6-305')] <= 1
                model += x_as[al, salas['Sala'].to_list().index('6-305/6-306')] + \
                x_as[a, salas['Sala'].to_list().index('6-306')] <= 1


# Para cada turma/disciplina 't'.
for t in range(lenT):
    # Para cada sala 's'.
    for s in range(lenS):
        # Adiciono uma restrição que verifica se a sala 's' foi utilizada pela turma/disciplina 't'.
        # Isso é feito através do valor atribuído de c_st, pois, quando ela possui o valor 0, a somatória da esquerda precisa ser
        # obrigatoriamente 0, ou seja, eu garanto que a sala 's' não foi utilizada pela turma/disciplina 't'.
        # Caso ela tenha o valor 1, então há espaço para que uma aula de 't' seja alocada na sala 's', e eu levo em conta o caso de
        # mais de uma aula de 't' ser alocada na mesma sala, até o número real de aulas da turma/disciplina 't'.
        model += xsum(x_as[a, s] for a in A_tt[t]) <= len(A_tt[t]) * c_st[s, t]

    # Com as restrições de verificação concluídas e adicionadas, coloco uma restrição que garante que o valor de y_t reflita
    # o número de salas diferentes usadas pela turma/disciplina 't'.
    model += y_t[t] >= xsum(c_st[s, t] for s in range(lenS))


# Restrição de superlotação de uma sala.
# Para cada aula 'a'.
if sys.argv[5] and sys.argv[6]:
    for a in range(len(A)):
        # Para cada sala 's'.
        for s in range(len(S)):
            # Coloco uma restrição que verifica se uma aula ocupa mais de uma porcentagem da sala.
            # Isso é verificado quando o lado esquerdo, que descreve a ocupação total de uma aula 'a' em uma sala 's',
            # ultrapassa o valor de alpha * cap_s[s], que descreve a porcentagem ocupada. Se o valor for ultrapassado,
            # a variável z_as recebe o valor 1, portanto o lado direito passa a ter um valor a mais que balanceia a desigualdade.
            # Ex: Suponha que alpha seja 0.85, que a aula 'a' tenha 40 alunos e a sala 's' possua 45 lugares.
            # 85% de 45 é 38.25, que é menor que 40, logo, z_as recebe o valor 1 e o lado direito passa a ser 38.25 + 45,
            # mantendo a restrição verdadeira, mas o valor de z_as será contabilizado na função objetivo e uma penalidade será aplicada.
            # Por conta dessa penalidade, o modelo tentará evitar fazer esse tipo de alocação.
            model += tam_t[int((a % lenT ))] * x_as[a, s] <= alpha * cap_s[s] + z_as[a, s] * cap_s[s]

if sys.argv[4]:
    # (3.13) - Relação entre w_cs e x_as
    # Restrição que estabelece a relação entre w_cs e x_as, isto é, que identifica se o curso 'c' terá uma aula alocada na sala 's'.
    # Para cada sala 's'.
    for s in range(lenS):
        # Para cada curso 'c'.
        for c in range(lenC):
            # Para cada aula do curso 'c'.
            for a in A_c[c]:
                # Adiciono a restrição que relaciona as duas variáveis. Como 'a' é uma aula ministrada para o curso 'c',
                # quando x_as = 1, então o curso 'c' possui uma aula na sala 's', e a variável w_cs é obrigada a ter o valor 1 neste caso.
                model += x_as[a,s] <= w_cs[c,s]
    
    # (3.14) - Ligação das variáveis w_cs e vcsl
    # Restrição que estabelece a relação entre w_cs e v_cssl, isto é, que identifica se o curso 'c' terá uma aula
    # alocada na sala 's' e outra na sala 'sl'.
    # Para cada curso 'c'.
    for c in range(lenC):
        # Para cada sala 's'.
        for s in range(lenS):
            # Para cada sala 'sl'
            for sl in range(lenS):
                # Verifico se 's' e 'sl' são a mesma sala.
                if s != sl:
                    # Caso não sejam, adiciono as restrições de relação. Quando w_cs = 1 e w_csl = 1, a segunda restrição garante que
                    # v_cssl ganhe o valor 1. No caso de que ao menos uma entre w_cs e w_csl tem o valor 0, a primeira restrição
                    # garante que v_cssl também seja 0.
                    # Em outras palavras, as restrições garantem que v_cssl = 1 apenas quando w_cs = 1 e w_csl = 1, que
                    # significa que o curso 'c' possui ao menos uma aula na sala 's' e uma na sala 'sl'.
                    model += 2 * v_cssl[c,s,sl] <= w_cs[c,s] + w_cs[c,sl]
                    model += v_cssl[c,s,sl] >= w_cs[c,s] + w_cs[c,sl] - 1

# Defino a tolerância do ótimo com algum valor, isto é, defino uma distância máxima que eu tolero entre uma solução factível e uma solução ótima.
model.opt_tol = 0.1
# Desabilito as saídas de execução do modelo.
model.verbose = 0
model.tol = 1e-6
model.integer_tol = 1e-6
# Começo a cronometrar o tempo gasto na execução do modelo.
start = time.time()
# O modelo começa a fazer a alocação, com tempo limite de 7200 segundos (2 horas) de tempo limite.
model.optimize(max_seconds=7200)
# Terminado a execução do modelo, paro o cronômetro e calculo o tempo utilizado.
end = time.time()
exec = end - start
print(f"Tempo de execução: {exec}")

# Escrevo qual o tipo de solução obtida, como ótima, factível, infactível, etc.
print(f"Status do modelo: {model.status}")

# A variável aulas serve para contar quantas aulas foram alocadas pelo modelo. Note que as aulas do LEM não são contabilizadas aqui.
aulas = 0
# A variável ocupacao serve para calcular a ocupação total de todas as aulas alocadas.
ocupacao = 0
# Para cada aula 'a'.
for a in range(lenA):
    # Para cada sala 's'.
    for s in range(lenS):
        # Se o valor de x_as for maior que 0.5, quer dizer que a aula 'a' foi alocada na sala 's'.
        if x_as[a,s].x >= 0.5:
            # Contabilizo ela na variável aulas, e adiciono a ocupação da aula 'a' na sala 's'
            aulas += 1
            ocupacao += uso_as[a,s]

# A variável trocas contabiliza as escolhas do modelo à respeito da variável y_t.
trocas = 0
# Para cada turma/disciplina 't'.
for t in range(lenT):
    # Se o valor de y_t for maior que 0.5, quer dizer que a turma/disciplina:
    # 1) teve uma ou mais trocas de sala, se a interpretação de y_t for o número de trocas de sala feitas pela turma/disciplina 't'.
    # 2) teve uma ou mais salas utilizadas, se a interpretação de y_t for o número de salas utilizadas pela turma/disciplina 't'.
    if y_t[t].x >= 1.5:
        # Imprimo o valor de y_t e contabilizo o ocorrido na variável trocas.
        trocas += 1


print(f"Valor da solução: {model.objective_value}")
print(f"Valor do GAP da solução: {model.gap:.2%}")
print(f"Valor das salas ocupadas: {ocupacao}.")

"""## Salvamento dos dados"""

# Com o modelo executado e a solução obtida, vou passar
# a salvar os dados de maneiras específicas, fazer certas verificações, visualizações e relatórios com os dados obtidos.

# Crio uma variável para conter o diretório da pasta que irá conter os arquivos gerados.
pasta_dados = os.path.join(os.getcwd(), "Saídas da Interface", "Saídas do Modelo")

# Crio listas para guardar os dados importantes das aulas, como o código da disciplina, o horário da aula, o número de inscritos, etc.
codigos = []
horarios = []
inscritos = []
salas_alocadas = []
nomes = []
docentes = []
curso_aula = []
ano_dados = []
nusp = []

# Para cada aula 'a'.
for a in range(lenA):
    # Salvo o índice da turma/disciplina dela para diminuir o número de ações do código.
    index = int(a % (lenT))
    # Também identifico qual coluna de horários do dataframe corresponde a aula 'a', isto é, qual coluna de horário 'a' pertence.
    coluna_horario = "Horário" + " " + str(1 + int(a / lenT))
    # Se a aula 'a' é ministrada no LEM e possui um horário definido
    if df.loc[index, 'Sala'] == '6-307' and not (pd.isna(df.loc[index, coluna_horario])):
        # Salvo o código da disciplina que 'a' pertence.
        codigos.append(df['Disciplina (código)'][index]) # deveria apendar apenas os códigos de matérias com horário ok
        # Salvo o horário da aula.
        horarios.append(df[coluna_horario][index])
        # Salvo o nome completo da disciplina que 'a' pertence.
        nomes.append(df['Disciplina (nome completo)'][index])
        # Salvo o nome do docente que irá ministrar a disciplina que 'a' pertence.
        docentes.append(df['Docente (nome completo sem abreviações)'][index])
        # Salvo o número de inscritos da disciplina que 'a' pertence.
        inscritos.append(df['Vagas por disciplina'][index])
        # Salvo a sala onde a aula 'a' foi alocada.
        salas_alocadas.append('6-307')
        # Salvo os cursos para os quais a disciplina que 'a' pertence é ministrada.
        curso_aula.append(df['Curso(s)'][index])
        nusp.append(df['NUSP'][index])
        ano_dados.append(df['Ano dos dados'][index])

    else:
        # Para cada sala 's'.
        for s in range(lenS):
            # Verifico se 'a' foi alocada em 's'.
            if x_as[a,s].x >= 0.5:
                # Salvo o código da disciplina que 'a' pertence.
                codigos.append(df['Disciplina (código)'][index]) # deveria apendar apenas os códigos de matérias com horário ok
                # Salvo o horário da aula.
                horarios.append(df[coluna_horario][index])
                # Salvo o nome completo da disciplina que 'a' pertence.
                nomes.append(df['Disciplina (nome completo)'][index])
                # Salvo o nome do docente que irá ministrar a disciplina que 'a' pertence.
                docentes.append(df['Docente (nome completo sem abreviações)'][index])
                # Salvo o número de inscritos da disciplina que 'a' pertence.
                inscritos.append(df['Vagas por disciplina'][index])
                # Salvo a sala onde a aula 'a' foi alocada.
                salas_alocadas.append(salas['Sala'][s])
                # Salvo os cursos para os quais a disciplina que 'a' pertence é ministrada.
                curso_aula.append(df['Curso(s)'][index])
                nusp.append(df['NUSP'][index])
                ano_dados.append(df['Ano dos dados'][index])


# A variável dados_solucao é um dicionário contendo as listas que acabamos de criar, e essa variável será
# usada para criar um novo dataframe que será salvo no formato de uma planilha do excel.
dados_solucao = {
    'Disciplina': codigos,
    'Nomes': nomes,
    'Cursos': curso_aula,
    'Horário': horarios,
    'Sala': salas_alocadas,
    'Inscritos': inscritos,
    'Docentes': docentes,
    'NUSP': nusp,
    'Ano dos dados': ano_dados,
    'Observações': [pd.NA for _ in range(len(codigos))]
}

# Converte os dados para um DataFrame.
dataframe = pd.DataFrame(dados_solucao)

# Caminho do novo arquivo Excel a ser criado.
# Esta planilha é uma visualização técnica de como as aulas foram alocadas, mostrando a disciplina em questão, o horário da aula daquela
# turma/disciplina. Note que as repetições aparentes das linhas se dá por conta dos diferentes horários das turmas/disciplinas.
full_name = "Dados da solução do Modelo.xlsx"
file_path = os.path.join(pasta_dados, full_name)


try:
    # Crio um novo arquivo Excel e escrevo os dados.
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        dataframe.to_excel(writer, sheet_name='Resultados', index=False)

    print(f"Novo arquivo '{full_name}' criado e dados salvos com sucesso!")
except PermissionError as e:
        if e.errno == 13:
            print(f"O arquivo '{full_name}' está aberto em algum programa (como o Excel). Feche o arquivo e tente novamente.")
            traceback.print_exc(file=sys.stderr)
            sys.exit(2)
        else:
            print(f"Ocorreu um erro inesperado ao criar o arquivo '{full_name}': {e}")
            traceback.print_exc(file=sys.stderr)
            sys.exit(1)
except Exception as e:
    print(f"Ocorreu um erro inesperado ao criar o arquivo '{full_name}': {e}")
    traceback.print_exc(file=sys.stderr)
    sys.exit(1)

"""## Factibilidade e Verificação"""

# Via de regra, se o modelo for infactível, é porque não tem como alocar todas as aulas nas salas.
# Com isso em mente, se em um primeiro momento o modelo rodar e retornar uma solução possível, mas após
# uma mudança de horário de uma disciplina/adição de turma não houver retorno,
# então as mudanças NÃO PODEM acontecer.

# O modelo deve retornar apenas soluções factíveis, já que as restrições são bem colocadas
# Para fazer uma melhor verificação, vamos montar um código que examina a saída e aponta quais disciplinas
# não foram alocadas. Estes casos provavelmente são referentes a erros de horários que não foram computados.
# Importante notar que a saída envolve disciplinas com 0 alunos, mas se uma disciplina não tem alunos, ela
# não deveria estar na planilha, afinal, não tem motivo para alocar uma sala para 0 pessoas.

# Traço o caminho até a planilha feita para o STI.
planilha_dados = pd.read_excel(os.path.join(pasta_dados, "Dados da solução do Modelo.xlsx"))

# Crio uma variável no formato de lista com todas as turmas/disciplinas cujas aulas foram alocadas.
alocadas = planilha_dados['Disciplina'].tolist()
# Crio uma variável no formato de lista com todas as turmas/disciplinas que deveriam ser alocadas.
disciplinas = df['Disciplina (código)'].tolist()

# Para cada turma/disciplina 't' na lista de turmas/disciplinas que deveriam ser alocadas.
for t in disciplinas:
    # Se ela não fazer parte da lista de turmas/disciplinas alocadas pelo modelo, escrevo um aviso para o usuário.
    if t not in alocadas:
        print(f"A disciplina {t} não foi alocada. Verifique os horários e o número de vagas, pode haver algum erro de digitação.")

"""## Planilha de Distribuição"""

# Cria uma paleta de cores para utilizar como preenchimento.
cores = [
    "FF5733",  # Laranja brilhante
    "33FF57",  # Verde brilhante
    "3357FF",  # Azul forte
    "F3FF33",  # Amarelo
    "FF33A1",  # Rosa forte
    "FF8C33",  # Laranja queimado
    "A133FF",  # Roxo
    "33FFF3",  # Ciano

]

# A variável dict é um dicionário na forma de matriz que é inicialmente repleta de 0's.
# Cada valor dela identifica o número de aulas em uma determinada sala que um determinado curso tem, ou seja,
# para cada curso 'c' e sala 's', o par (c,s) mostra o número de aulas que o curso 'c' possui em 's'.
dict = {(c,s): 0 for c in range(lenC) for s in range(lenS)}

# A variável naulas_s é uma lista que contém o número de aulas total de cada sala.
naulas_s = []

# Para cada sala 's'.
for s in range(lenS):
    # Defino uma variável que conta o número de aulas de cada sala.
    naulas = 0
    # Para cada aula 'a'.
    for a in range(lenA):
        # Verifico se a aula 'a' foi alocada na sala 's'.
        if x_as[a,s].x >= 0.5:
            # Em caso positivo, obtenho a turma/disciplina para a qual a aula 'a' pertence.
            t = int(a % lenT)
            # Contabilizo a aula na variável contadora.
            naulas += 1
            # Para curso 'c'.
            for c in range(lenC):
                # Atualizo o valor de dict correspondente a contagem de aulas que o curso 'c' possui na sala 's'.
                dict.update({(c,s): dict[c,s] + Y_tc[t,c]})

    # Adiciono o número de aulas alocadas na sala 's' na lista naulas_s.
    naulas_s.append(naulas)

# A variável curso é uma lista contendo os índices dos cursos do ICMC.
curso = [t[0] for t in list(dict.keys())]
# A variável room é uma lista contendo os índices dos salas do ICMC.
room = [t[1] for t in list(dict.keys())]


# Cria o workbook, isto é, um objeto de planilha do Excel.
wb = Workbook()
ws = wb.active
ws.title = "Distribuição de Cursos"

# Começo a montara tabela colocando o nome das salas na primeira coluna da planilha.
for row in room:
    ws.cell(row=row+2,column=1).value = salas.loc[row, 'Sala']

# Feito isso, adiciono o nome dos cursos do ICMC na primeira linha da planilha.
for col in curso:
    ws.cell(row=1,column=col+2).value = curriculos[col]

# Para cada curso na lista de cursos, que serão vistos como as colunas da planilha.
for col in curso:
    # Para cada sala na lista de salas, que serão vistas como as linhas da planilha.
    for row in room:
        # Defino o valor da célula na linha e coluna atual (sala e curso atual, respectivamente) como o valor equivalente de dict.
        ws.cell(row=row+2,column=col+2).value = dict[col,row]

        # Defino o valor mínimo de uma barra de progresso como 0.
        min_obj = FormatObject(type="num", val=0)

        # Verifico se o número de aulas alocadas na sala 'row' é 0.
        if naulas_s[row] == 0:
            # Se for, defino o valor máximo de uma barra de progresso como 1, apenas para visualização.
            max_obj = FormatObject(type="num", val=1)
        else:
            # Se não for, defino o valor máximo de uma barra de progresso como o esse número.
            max_obj = FormatObject(type="num", val=naulas_s[row])

        # A variável cor recebe uma das cores da paleta criada anteriormente. O cálculo para determinar isso não é muito importante,
        # ele só foi feito assim para garantir que as cores não se repetiriam muito próximas umas das outras.
        cor = cores[int(row % lenC)]

        # Defino, então, a barra de progresso, utilizando os valor mínimo e máximo definidos anteriormente, e definimos a cor da mesma.
        data_bar = DataBar(cfvo=[min_obj, max_obj], color=cor, showValue="None")

        # Defino uma regra para colocar em uma linha da planilha, representando uma sala.
        rule = Rule(type="dataBar", dataBar=data_bar)

        # Adiciona a barra de progresso na linha correspondente.
        # Note que a "linha" é considerada como a célula da segunda coluna (coluna B) da planilha, até a célula da mesma linha
        # com a coluna sendo a do último curso da lista de cursos.
        ws.conditional_formatting.add(f"B{str(row+2)}:{chr(64+lenC+1)}{str(row+2)}", rule)

# Com as barras de progresso no lugar certo, adiciono uma coluna com o número de aulas de cada sala após a coluna do último curso.
ws.cell(row=1,column=int(len(curso)/lenS)+2).value = 'Nº Aulas'
# Para cada sala na lista de salas.
for row in room:
    # Coloco o número de aulas alocadas naquela sala.
    ws.cell(row=row+2,column=int(len(curso)/lenS)+2).value = naulas_s[row]

# Logo abaixo da última sala adicionada, eu coloco o valor da ocupação total e do tempo de execução do modelo em segundos.
ws.cell(row=lenS+2,column=1).value = ocupacao
ws.cell(row=lenS+3,column=1).value = exec

# Por fim, salvo o arquivo.
full_name = f"Distribuição de Cursos.xlsx"

file_path = os.path.join(pasta_dados, full_name)

try:
    wb.save(file_path)
    print(f"Arquivo '{full_name}' salvo com sucesso!")
except PermissionError as e:
    if e.errno == 13:
        print(f"O arquivo '{full_name}' está aberto em algum programa (como o Excel). Feche o arquivo e tente novamente.")
        traceback.print_exc(file=sys.stderr)
        sys.exit(2)
    else:
        print(f"Ocorreu um erro inesperado ao criar o arquivo '{full_name}': {e}")
        traceback.print_exc(file=sys.stderr)
        sys.exit(1)
except Exception as e:
    print(f"Ocorreu um erro inesperado ao criar o arquivo '{full_name}': {e}")
    traceback.print_exc(file=sys.stderr)
    sys.exit(1)

